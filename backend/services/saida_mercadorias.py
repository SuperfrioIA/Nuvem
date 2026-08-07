"""Leitura controlada da familia SAIDA_MERCADORIAS do SharePoint DataHub
(Lote V2.3).

Mesma guarda de seguranca do leitor de entrada (item_id so aceito se estiver
na ultima sincronizacao do inventario), mas quatro diferencas de fundo, todas
conferidas no dado real em 06/ago/2026 (docs/V2_3_PLANO_EXECUCAO.md):

1. **Cabecalho em DOIS niveis.** Linha 5 tem as seis bandas (GSM, Produto,
   Solicitado pelo Cliente, Atendido pelo Estoque, Separado Fisicamente, Dados
   de Separacao); linha 6 tem os 34/36 rotulos reais, com seis deles repetidos
   3x (Volume, EMB, Fracao, EMB, Peso Liquido, Peso Bruto -- uma vez por
   banda de medida). Valida-se as DUAS linhas: so a linha 6 nao distingue os
   dois layouts (os rotulos sao identicos, so a posicao muda); so a linha 5
   nao pega banda com rotulo interno fora de ordem.
2. **A posicao do Peso Bruto NAO e fixa.** Depende do layout: coluna 31 nas
   unidades com 36 colunas (RMSPII, CWB3, RJ), coluna 29 na SANCA (34
   colunas, sem Cliente/Cliente CNPJ). O leitor acha a banda "Separado
   Fisicamente" pela LINHA 5 e le "Peso Bruto" no deslocamento fixo (+5)
   dentro dela -- nunca posicao chumbada. Ler a coluna 31 num arquivo da
   SANCA leria "Inicio" (um timestamp) como peso.
3. **Nao existe coluna de valor**, em nenhuma unidade -- so peso e contagem
   (decisao D1 do V2.3: nao existe metrica peso_mercadoria_saida... digo,
   valor_mercadoria_saida).
4. **Agregacao em streaming.** Um arquivo real chega a 99628 linhas de dado
   (SAIDA_MERCADORIAS_025_2607_f1). Materializar tudo numa lista, como o
   leitor de entrada faz (arquivos de ~400 KB, sem esse problema), nao
   escala aqui. `ler()` devolve um GERADOR de linhas normalizadas; os
   contadores (lidas/validas/descartadas/canceladas) so ficam definitivos
   depois que o gerador for consumido por completo -- e a agregacao do
   processamento consome exatamente uma vez, sem materializar a lista.

Arquivos vem partidos em `_f1`/`_f2` (33 MB por filial/competencia) OU sem
sufixo nenhum (a CWB3 publica assim) -- `dados_da_familia` devolve o indice da
parte (ou None pra "sem sufixo"), e quem decide o que fazer com varias partes
da mesma competencia e o processamento (backend/services/processamento_datahub.py),
nao este modulo: aqui cada `ler()` cobre UM arquivo/UMA parte.
"""

import io
import os
import re
import unicodedata

import openpyxl

from . import graph_datahub, inventario_datahub

_ABA_ESPERADA = "SLIN"

_PADRAO_NOME = re.compile(
    r"^SAIDA_MERCADORIAS_(\d+(?:-\d+)*)_(\d{2})(\d{2})(?:_f(\d+))?\.xlsx$", re.IGNORECASE
)

# As seis bandas da linha 5, na ordem esperada (posicoes absolutas variam por
# layout -- so a ORDEM relativa e fixa). Normalizadas (sem acento, maiusculas)
# pra comparar sem depender de como o Excel/openpyxl devolveu o acento.
_SEQUENCIA_BANDAS = (
    "GSM", "PRODUTO", "SOLICITADO PELO CLIENTE", "ATENDIDO PELO ESTOQUE",
    "SEPARADO FISICAMENTE", "DADOS DE SEPARACAO",
)
_BANDA_OFICIAL = "SEPARADO FISICAMENTE"

# Os seis rotulos de UMA banda de medida, na ordem esperada dentro dela --
# repetem 3x na linha 6 (um conjunto por banda). Peso Bruto e o ULTIMO da
# banda oficial: deslocamento +5 a partir do inicio dela.
_ROTULOS_BANDA_MEDIDA = ("VOLUME", "EMB", "FRACAO", "EMB", "PESO LIQUIDO", "PESO BRUTO")
_DESLOCAMENTO_PESO_BRUTO = 5

# Rotulos unicos (nao repetidos) que tem que existir nos DOIS layouts, fora
# das tres bandas de medida e fora de Cliente/Cliente CNPJ (unicos opcionais).
_ROTULOS_SEMPRE = (
    "Estoque", "Empresa", "GSM", "Operação", "Data Solicitação", "Data Saída",
    "Status Separação", "Item", "Código", "Descrição", "Pedido", "Destinatário",
    "Corte Físico", "Início", "Final", "Separador",
)
_COLUNAS_CLIENTE = ("Cliente", "Cliente CNPJ")

LAYOUT_36_COLUNAS = "36_colunas"
LAYOUT_34_COLUNAS = "34_colunas"

_STATUS_CANCELADO = "CANCELADO"


class SaidaMercadoriasError(Exception):
    """Erro de validacao do item/arquivo/aba/banda/coluna -- mensagem clara
    pro chamador (endpoint/script traduz pra HTTP 400 ou saida com erro)."""


def _limite_bytes() -> int:
    limite_mb = int(os.environ.get("UPLOAD_MAX_MB", "50"))
    return limite_mb * 1024 * 1024


def _normalizar(valor) -> str:
    texto = str(valor if valor is not None else "").strip().upper()
    sem_acento = unicodedata.normalize("NFKD", texto).encode("ascii", "ignore").decode("ascii")
    return sem_acento


def arquivo_do_inventario(item_id: str) -> dict:
    if not inventario_datahub.status().get("resumo"):
        raise SaidaMercadoriasError(
            "nenhuma sincronizacao do DataHub ainda -- clique em 'Sincronizar agora' primeiro"
        )
    arquivo = inventario_datahub.arquivo_por_item_id(item_id)
    if arquivo is None:
        raise SaidaMercadoriasError(
            "item_id nao encontrado na ultima sincronizacao do DataHub"
        )
    return arquivo


def dados_da_familia(nome) -> tuple[str, str, int | None] | None:
    """(filial, competencia, indice_parte) quando o nome pertence a familia
    SAIDA_MERCADORIAS; None caso contrario. `indice_parte` e None quando o
    arquivo nao tem sufixo `_fN` (a CWB3 publica assim -- parte unica,
    indiferenciada) ou o inteiro do sufixo quando tem (`_f1`, `_f2`, ...).
    """
    m = _PADRAO_NOME.match(nome or "")
    if not m:
        return None
    filial, aa, mm, parte = m.group(1), m.group(2), m.group(3), m.group(4)
    indice_parte = int(parte) if parte is not None else None
    return filial, f"20{aa}-{mm}", indice_parte


def _validar_nome(nome: str) -> tuple[str, str, int | None]:
    if not nome.lower().endswith(".xlsx"):
        raise SaidaMercadoriasError(f"extensao invalida (esperado .xlsx): {nome}")
    dados = dados_da_familia(nome)
    if dados is None:
        raise SaidaMercadoriasError(
            "nome de arquivo fora do padrao "
            f"SAIDA_MERCADORIAS_{{filial}}_{{AAMM}}[_f{{N}}].xlsx: {nome}"
        )
    return dados


def _abrir_aba(conteudo: bytes):
    try:
        wb = openpyxl.load_workbook(io.BytesIO(conteudo), read_only=True, data_only=True)
    except Exception as exc:
        raise SaidaMercadoriasError("arquivo nao e um .xlsx valido ou esta corrompido") from exc
    if _ABA_ESPERADA not in wb.sheetnames:
        wb.close()
        raise SaidaMercadoriasError(f"aba '{_ABA_ESPERADA}' nao encontrada no arquivo")
    return wb


def _bandas(linha5) -> dict[str, int]:
    bandas: dict[str, int] = {}
    for i, valor in enumerate(linha5 or ()):
        nome = _normalizar(valor)
        if nome and nome not in bandas:
            bandas[nome] = i
    return bandas


def _posicao_banda_oficial(bandas: dict[str, int]) -> int:
    faltando = [b for b in _SEQUENCIA_BANDAS if b not in bandas]
    if faltando:
        raise SaidaMercadoriasError(
            "banda(s) nao encontrada(s) na linha 5 (cabecalho de dois niveis): "
            + ", ".join(faltando)
        )
    posicoes = [bandas[b] for b in _SEQUENCIA_BANDAS]
    if posicoes != sorted(posicoes):
        raise SaidaMercadoriasError(
            "bandas da linha 5 fora da ordem esperada "
            f"({_SEQUENCIA_BANDAS}) -- cabecalho reordenado ou invalido"
        )
    return bandas[_BANDA_OFICIAL]


def _validar_banda_oficial(linha6, inicio: int) -> None:
    """Confere que os 6 rotulos a partir do inicio da banda oficial batem
    EXATAMENTE com Volume/EMB/Fracao/EMB/Peso Liquido/Peso Bruto, nessa
    ordem -- e so essa validacao que garante que o deslocamento +5 aponta
    pra Peso Bruto de verdade, e nao pra outra coisa por banda deslocada."""
    fim = inicio + len(_ROTULOS_BANDA_MEDIDA)
    if fim > len(linha6):
        raise SaidaMercadoriasError(
            "linha 6 mais curta que o esperado pra banda 'Separado Fisicamente'"
        )
    encontrados = tuple(_normalizar(v) for v in linha6[inicio:fim])
    if encontrados != _ROTULOS_BANDA_MEDIDA:
        raise SaidaMercadoriasError(
            "rotulos da banda 'Separado Fisicamente' (linha 6) nao batem com o "
            f"esperado {_ROTULOS_BANDA_MEDIDA} -- encontrado {encontrados}"
        )


def _indice_rotulos(linha6) -> dict[str, int]:
    """Indice nome->posicao dos rotulos UNICOS (fora das tres bandas de
    medida, que repetem e por isso NUNCA sao lidos por nome -- so por
    deslocamento a partir da banda oficial, ver `_validar_banda_oficial`)."""
    indice: dict[str, int] = {}
    for i, valor in enumerate(linha6 or ()):
        nome = str(valor).strip() if valor is not None else ""
        if nome and nome not in indice:
            indice[nome] = i
    faltando = sorted({c for c in _ROTULOS_SEMPRE if c not in indice})
    if faltando:
        raise SaidaMercadoriasError("coluna(s) nao encontrada(s): " + ", ".join(faltando))
    return indice


def _detectar_layout(cabecalho5, cabecalho6) -> tuple[dict[str, int], int, str]:
    """(indice_rotulos_unicos, coluna_do_peso_bruto, layout). Valida a linha 5
    E a linha 6 -- ver docstring do modulo."""
    bandas = _bandas(cabecalho5)
    inicio_oficial = _posicao_banda_oficial(bandas)
    _validar_banda_oficial(cabecalho6, inicio_oficial)
    col_peso_bruto = inicio_oficial + _DESLOCAMENTO_PESO_BRUTO

    indice = _indice_rotulos(cabecalho6)
    presentes_cliente = [c for c in _COLUNAS_CLIENTE if c in indice]
    if presentes_cliente and len(presentes_cliente) != len(_COLUNAS_CLIENTE):
        raise SaidaMercadoriasError(
            "layout inconsistente: tem "
            f"{presentes_cliente[0]!r} mas nao {[c for c in _COLUNAS_CLIENTE if c not in presentes_cliente]!r}"
        )
    layout = LAYOUT_36_COLUNAS if presentes_cliente else LAYOUT_34_COLUNAS
    return indice, col_peso_bruto, layout


def ler(item_id: str) -> dict:
    """Baixa, valida o cabecalho de dois niveis e devolve metadados + um
    GERADOR de linhas normalizadas (nao materializa -- ver docstring do
    modulo). `linhas` so pode ser consumido UMA vez; `contadores` (dict
    mutavel) so fica definitivo depois de esgotar o gerador por completo --
    e assim que o chamador sabe `linhas_lidas`/`linhas_validas`/`sem_dado`
    depois de agregar."""
    arquivo = arquivo_do_inventario(item_id)
    nome = arquivo["nome"]
    filial, competencia, indice_parte = _validar_nome(nome)

    conteudo = graph_datahub.baixar_item(item_id, limite_bytes=_limite_bytes())
    wb = _abrir_aba(conteudo)
    ws = wb[_ABA_ESPERADA]
    it = ws.iter_rows(values_only=True)

    linha5 = linha6 = None
    for i in range(1, 7):
        linha = next(it, None)
        if linha is None:
            wb.close()
            raise SaidaMercadoriasError(
                f"arquivo com menos de 6 linhas -- sem cabecalho de dois niveis (parou na {i})"
            )
        if i == 5:
            linha5 = linha
        elif i == 6:
            linha6 = linha

    indice, col_peso_bruto, layout = _detectar_layout(linha5, linha6)
    col_cliente = indice.get("Cliente")
    col_cliente_cnpj = indice.get("Cliente CNPJ")
    col_estoque = indice["Estoque"]
    col_status = indice["Status Separação"]

    contadores = {"lidas": 0, "validas": 0, "descartadas": 0, "canceladas": 0}

    def _linhas():
        try:
            for linha in it:
                if linha is None or all(v is None for v in linha):
                    continue
                contadores["lidas"] += 1

                status = _normalizar(linha[col_status] if col_status < len(linha) else None)
                if status == _STATUS_CANCELADO:
                    contadores["canceladas"] += 1
                    continue

                peso_bruto = _paranum_br(
                    linha[col_peso_bruto] if col_peso_bruto < len(linha) else None
                )
                if peso_bruto is None:
                    contadores["descartadas"] += 1
                    continue

                contadores["validas"] += 1
                yield {
                    "Cliente": (
                        linha[col_cliente] if col_cliente is not None and col_cliente < len(linha)
                        else None
                    ),
                    "Cliente CNPJ": (
                        linha[col_cliente_cnpj]
                        if col_cliente_cnpj is not None and col_cliente_cnpj < len(linha)
                        else None
                    ),
                    # nome canonico "Nome Estoque" (nao "Estoque"): e o campo
                    # que tipo_estoque.classificar() e a agregacao esperam --
                    # mesma chave da familia de entrada, pra reaproveitar o
                    # agregador sem duplicar logica (processamento_datahub.py)
                    "Nome Estoque": linha[col_estoque] if col_estoque < len(linha) else None,
                    "Peso Bruto": peso_bruto,
                }
        finally:
            wb.close()

    return {
        "arquivo": nome,
        "caminho": arquivo.get("caminho"),
        "modificado_em": arquivo.get("modificado_em"),
        "tamanho": arquivo.get("tamanho"),
        "unidade": inventario_datahub.unidade_do_caminho(arquivo.get("caminho")),
        "filial": filial,
        "competencia": competencia,
        "indice_parte": indice_parte,
        "layout": layout,
        "linhas": _linhas(),
        "contadores": contadores,
    }


def _paranum_br(valor):
    if valor is None:
        return None
    if isinstance(valor, (int, float)):
        return float(valor)
    texto = str(valor).strip()
    if not texto:
        return None
    if "," in texto:
        texto = texto.replace(".", "").replace(",", ".")
    try:
        return float(texto)
    except ValueError:
        return None
