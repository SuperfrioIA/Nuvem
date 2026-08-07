"""Leitura controlada da familia ENTRADA_MERCADORIAS do SharePoint DataHub
(Lote P3).

Nao e um leitor generico de planilha -- so entende esta familia especifica
(nome, aba e cabecalho fixos, documentados em docs/FONTES_DATAHUB.md). O
caminho generico de leitura continua sendo o dos modelos de importacao
(backend/conectores/upload_manual.py).

**Dois layouts (V2.3).** A RJ (`RJ/004-003`) publica uma variante de 18
colunas, sem `Cliente` e `Cliente CNPJ` -- conferido no dado em 06/ago/2026
(docs/V2_3_PLANO_EXECUCAO.md). O layout e detectado pelo CABECALHO, nunca pela
unidade: um de-para novo nao pode mudar como um arquivo e lido, e amanha outra
unidade pode publicar a mesma variante. No layout de 18, toda linha entra no
balde `cliente_id NULL` (decisao D2 -- nao ha CNPJ pra cadastrar, entao
`raiz_cnpj(None)` ja devolve None e `processamento_datahub` ja NAO registra
pendencia de cliente quando a raiz e None -- nenhum codigo novo precisou disso,
so este leitor emitir `None` em vez de faltar a chave).

Guarda de seguranca: item_id so e aceito se aparecer na lista de arquivos da
ultima sincronizacao do Lote P2 (backend/services/inventario_datahub.py) --
nunca um id digitado ou uma URL externa. Alem disso, o nome do arquivo tem que
bater no padrao ENTRADA_MERCADORIAS_{filial}_{AAMM}.xlsx, a extensao tem que
ser .xlsx e a aba tem que se chamar SLIN -- qualquer uma falhando, erro claro
(EntradaMercadoriasError).

openpyxl nao e o Excel e nao executa macro -- read_only=True e data_only=True
atendem a exigencia de "sem macro" por construcao, sem checagem extra.

Colunas sao localizadas por nome (dicionario cabecalho->indice montado na
leitura), nunca por posicao chumbada -- se a planilha ganhar uma coluna nova
amanha, a leitura continua funcionando. Posicao so seria necessaria pra
desempatar o rotulo duplicado (EMB aparece duas vezes), mas nenhuma coluna
candidata a KPI usa EMB, entao a primeira ocorrencia basta.
"""

import io
import os
import re

import openpyxl

from . import filiais_datahub, graph_datahub, inventario_datahub

_ABA_ESPERADA = "SLIN"

# O codigo de filial aceita hifen (`004-003`, `005-001`): e como a unidade RJ
# nomeia os exports dela. Antes o padrao exigia so digitos, entao os 42
# arquivos da RJ nao casavam e sumiam do processamento SEM virar pendencia --
# "nao casou no regex" virava "nao existe". Casando, eles chegam ate a
# resolucao de de-para (V2.1: parava ali, visivel como pendencia, porque a RJ
# nao tinha de-para nem leitor da variante; V2.3: de-para e leitor da
# variante de 18 colunas entram juntos, ver docstring do modulo).
_PADRAO_NOME = re.compile(
    r"^ENTRADA_MERCADORIAS_(\d+(?:-\d+)*)_(\d{2})(\d{2})\.xlsx$", re.IGNORECASE
)

# As 20 colunas do export real (docs/FONTES_DATAHUB.md) -- todas obrigatorias
# no cabecalho (decisao de 29/jul/2026: validar todas, nao so as dos KPIs do
# P4). EMB aparece duas vezes de proposito (posicoes 10 e 12 no arquivo real).
_COLUNAS_ESPERADAS = (
    "Cliente", "Cliente CNPJ", "GEM", "Devolução", "Solicitação", "NF Entrada",
    "Código", "Descrição", "Volume", "EMB", "Fração", "EMB", "Peso Líquido",
    "Peso Bruto", "Vlr. Unitário", "Vlr. Total", "Qtde UA", "Código Estoque",
    "Nome Estoque", "Operação",
)

# So estas recebem validacao de valor (parse numerico, linha descartada se
# falhar) -- as outras 14 sao lidas cruas, sem formato esperado definido
# (decisao de 29/jul/2026).
_COLUNAS_NUMERICAS = (
    "Volume", "Peso Líquido", "Peso Bruto", "Vlr. Unitário", "Vlr. Total", "Qtde UA",
)

# As duas colunas que a variante de 18 colunas da RJ nao tem (V2.3). Sao as
# unicas opcionais -- todas as outras 18 continuam obrigatorias nos dois
# layouts.
_COLUNAS_CLIENTE = ("Cliente", "Cliente CNPJ")
_COLUNAS_OBRIGATORIAS_SEMPRE = tuple(c for c in _COLUNAS_ESPERADAS if c not in _COLUNAS_CLIENTE)

LAYOUT_20_COLUNAS = "20_colunas"
LAYOUT_18_COLUNAS = "18_colunas"


class EntradaMercadoriasError(Exception):
    """Erro de validacao do item/arquivo/aba/coluna -- mensagem sempre clara
    pro chamador (endpoint traduz pra HTTP 400)."""


def _limite_bytes() -> int:
    """Mesmo limite do upload manual (UPLOAD_MAX_MB, default 50 MB) -- os
    arquivos desta familia tem ~400 KB, folga enorme; sem variavel nova."""
    limite_mb = int(os.environ.get("UPLOAD_MAX_MB", "50"))
    return limite_mb * 1024 * 1024


def arquivo_do_inventario(item_id: str) -> dict:
    """So aceita item_id que apareceu na ultima sincronizacao do P2 -- o cache
    vira lista de permissao (fonte unica: inventario_datahub.arquivo_por_item_id).
    Sem sincronizacao ainda, ou id desconhecido, falha com mensagem clara em vez
    de tentar baixar de qualquer forma."""
    if not inventario_datahub.status().get("resumo"):
        raise EntradaMercadoriasError(
            "nenhuma sincronizacao do DataHub ainda -- clique em 'Sincronizar agora' primeiro"
        )
    arquivo = inventario_datahub.arquivo_por_item_id(item_id)
    if arquivo is None:
        raise EntradaMercadoriasError(
            "item_id nao encontrado na ultima sincronizacao do DataHub"
        )
    return arquivo


def dados_da_familia(nome) -> tuple[str, str] | None:
    """(filial, competencia) quando o nome pertence a familia
    ENTRADA_MERCADORIAS; None caso contrario (V1.3 usa pra listar e rotular o
    que processar)."""
    m = _PADRAO_NOME.match(nome or "")
    if not m:
        return None
    filial, aa, mm = m.group(1), m.group(2), m.group(3)
    return filial, f"20{aa}-{mm}"


def _validar_nome(nome: str) -> tuple[str, str]:
    if not nome.lower().endswith(".xlsx"):
        raise EntradaMercadoriasError(f"extensao invalida (esperado .xlsx): {nome}")

    dados = dados_da_familia(nome)
    if dados is None:
        raise EntradaMercadoriasError(
            "nome de arquivo fora do padrao "
            f"ENTRADA_MERCADORIAS_{{filial}}_{{AAMM}}.xlsx: {nome}"
        )
    return dados


def _abrir_aba(conteudo: bytes):
    try:
        wb = openpyxl.load_workbook(io.BytesIO(conteudo), read_only=True, data_only=True)
    except Exception as exc:
        raise EntradaMercadoriasError("arquivo nao e um .xlsx valido ou esta corrompido") from exc

    if _ABA_ESPERADA not in wb.sheetnames:
        wb.close()
        raise EntradaMercadoriasError(f"aba '{_ABA_ESPERADA}' nao encontrada no arquivo")
    return wb


def _indice_cabecalho(linha_cabecalho) -> tuple[dict[str, int], str]:
    """Indice nome->posicao e o LAYOUT detectado (20 ou 18 colunas, V2.3).

    As 18 colunas sempre obrigatorias tem que existir nos dois layouts; so
    `Cliente`/`Cliente CNPJ` sao opcionais, e so as duas juntas -- ter uma sem
    a outra e planilha incoerente, nunca um layout valido, e falha com erro
    claro em vez de decidir por chute qual delas usar.
    """
    indice: dict[str, int] = {}
    for i, valor in enumerate(linha_cabecalho):
        nome = str(valor).strip() if valor is not None else ""
        if nome and nome not in indice:
            indice[nome] = i

    faltando = sorted({c for c in _COLUNAS_OBRIGATORIAS_SEMPRE if c not in indice})
    if faltando:
        raise EntradaMercadoriasError("coluna(s) nao encontrada(s): " + ", ".join(faltando))

    presentes_cliente = [c for c in _COLUNAS_CLIENTE if c in indice]
    if presentes_cliente and len(presentes_cliente) != len(_COLUNAS_CLIENTE):
        raise EntradaMercadoriasError(
            "layout inconsistente: tem "
            f"{presentes_cliente[0]!r} mas nao {[c for c in _COLUNAS_CLIENTE if c not in presentes_cliente]!r}"
        )
    layout = LAYOUT_20_COLUNAS if presentes_cliente else LAYOUT_18_COLUNAS
    return indice, layout


def _paranum_br(valor):
    """Aceita numero ja nativo (openpyxl le celula numerica como int/float) ou
    texto no formato BR (ponto de milhar, virgula decimal: '1.234,56').
    Devolve None se nao for nenhum dos dois -- linha e descartada."""
    if valor is None:
        return None
    if isinstance(valor, (int, float)):
        return float(valor)

    texto = str(valor).strip()
    if not texto:
        return None
    if "," in texto:
        texto = texto.replace(".", "").replace(",", ".")
    try:
        return float(texto)
    except ValueError:
        return None


def item_mais_recente() -> str:
    """item_id do arquivo ENTRADA_MERCADORIAS mais recente no inventario,
    RESTRITO a unidade representativa (filiais_datahub.UNIDADE_REPRESENTATIVA).

    Decisao de 29/jul/2026: a tela de KPIs (Lote P4) nao deixa escolher entre
    os arquivos da familia -- sempre mostra o mais recente sincronizado, mais
    simples pro POC.

    O recorte por unidade veio depois, com a reestruturacao da fonte: sem ele o
    "mais recente da familia" pode ser um arquivo de CWB3 ou da RJ, e a tela
    executiva passaria a exibir numero de outra unidade sob o rotulo da RMSPII
    (a CWB3 usa o mesmo codigo `001`), ou quebraria na leitura (a RJ tem 18
    colunas).

    Ate o V2.1 esse recorte era DERIVADO do mapa de de-para, e isso estava
    errado: dar de-para pra CWB3 e SANCA -- que era justamente o objetivo do
    lote -- expandiria o recorte de graca, e o card executivo passaria a poder
    exibir o arquivo mais recente de Curitiba sob o rotulo da RMSPII, sem
    ninguem pedir. Ter de-para e poder gravar; ser representativo e outra
    pergunta. Agora a unidade e explicita, e ampliar cobertura de ingestao nao
    mexe mais nesta tela.
    """
    resumo = inventario_datahub.status().get("resumo")
    if not resumo:
        raise EntradaMercadoriasError(
            "nenhuma sincronizacao do DataHub ainda -- clique em 'Sincronizar agora' primeiro"
        )
    unidade = filiais_datahub.UNIDADE_REPRESENTATIVA
    candidatos = [
        a for a in resumo.get("arquivos", [])
        if _PADRAO_NOME.match(a.get("nome", ""))
        and inventario_datahub.unidade_do_caminho(a.get("caminho")) == unidade
    ]
    if not candidatos:
        raise EntradaMercadoriasError(
            f"nenhum arquivo ENTRADA_MERCADORIAS da unidade {unidade} encontrado "
            "na ultima sincronizacao"
        )
    mais_recente = max(candidatos, key=lambda a: a.get("modificado_em") or "")
    return mais_recente["id"]


def ler(item_id: str) -> dict:
    """Baixa, valida e le a planilha; devolve metadados + linhas validadas.

    Todas as linhas validas ficam em memoria (o Lote P4 vai precisar delas
    pra somar os KPIs) -- quem expoe isso via HTTP decide quanto devolver.
    """
    arquivo = arquivo_do_inventario(item_id)
    nome = arquivo["nome"]
    filial, competencia = _validar_nome(nome)

    conteudo = graph_datahub.baixar_item(item_id, limite_bytes=_limite_bytes())

    wb = _abrir_aba(conteudo)
    try:
        ws = wb[_ABA_ESPERADA]
        linhas = ws.iter_rows(values_only=True)
        try:
            cabecalho = next(linhas)
        except StopIteration as exc:
            raise EntradaMercadoriasError("arquivo vazio -- sem linha de cabecalho") from exc

        indice, layout = _indice_cabecalho(cabecalho)
        colunas_unicas = list(dict.fromkeys(_COLUNAS_ESPERADAS))
        # No layout de 18, Cliente/Cliente CNPJ nao existem no cabecalho --
        # nao tem posicao pra ler. O registro recebe None pras duas explicito
        # (ver docstring do modulo: raiz_cnpj(None) ja cai no balde sem
        # cliente, sem pendencia, sem precisar de codigo novo no processamento).
        colunas_a_ler = [c for c in colunas_unicas if c in indice]

        linhas_validas: list[dict] = []
        lidas = 0
        descartadas = 0
        for linha in linhas:
            if linha is None or all(v is None for v in linha):
                continue
            lidas += 1

            registro = {c: None for c in _COLUNAS_CLIENTE}
            valida = True
            for coluna in colunas_a_ler:
                pos = indice[coluna]
                valor_bruto = linha[pos] if pos < len(linha) else None
                if coluna in _COLUNAS_NUMERICAS:
                    numero = _paranum_br(valor_bruto)
                    if numero is None:
                        valida = False
                        break
                    registro[coluna] = numero
                else:
                    registro[coluna] = valor_bruto

            if valida:
                linhas_validas.append(registro)
            else:
                descartadas += 1

    finally:
        wb.close()

    total_validas = len(linhas_validas)
    qualidade_pct = round(100 * total_validas / lidas, 1) if lidas else 0.0

    return {
        "arquivo": nome,
        "caminho": arquivo.get("caminho"),
        "modificado_em": arquivo.get("modificado_em"),
        "tamanho": arquivo.get("tamanho"),
        # unidade da fonte (galho de primeiro nivel): sozinho o codigo de
        # filial nao identifica armazem -- `001` existe em RMSPII e em CWB3
        "unidade": inventario_datahub.unidade_do_caminho(arquivo.get("caminho")),
        "filial": filial,
        "competencia": competencia,
        # layout detectado pelo cabecalho (V2.3) -- alimenta processamentos_datahub.layout_lido,
        # a base de "quais unidades nao tem coluna de cliente" (nunca uma lista escrita a mao)
        "layout": layout,
        "linhas_lidas": lidas,
        "linhas_validas": total_validas,
        "linhas_descartadas": descartadas,
        "qualidade_pct": qualidade_pct,
        # Cabecalho valido e ZERO linhas de dado: competencia sem movimento, um
        # estado legitimo da fonte (a SANCA comecou a operar em 2606, os arquivos
        # dela de 2601 a 2605 sao so cabecalho). Ate o V2.1.1 isto levantava
        # excecao e o processamento marcava `erro`, com dois efeitos ruins: cinco
        # erros permanentes no painel que ninguem pode resolver, e re-download em
        # toda rodada (o "pula inalterado" exige status `ok`).
        #
        # Quem decide o que fazer com isso e o CHAMADOR, nao o leitor: o
        # processamento grava status `sem_dado`; os endpoints que exibem UM
        # arquivo recusam com mensagem clara, pra tela executiva nunca renderizar
        # zero como se fosse medicao.
        "sem_dado": lidas == 0,
        "linhas": linhas_validas,
    }
