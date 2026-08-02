"""Leitura ESTRUTURAL generica de qualquer arquivo do inventario do DataHub
(Bloco D / V1.4 -- Laboratorio).

Nao substitui o leitor da familia integrada (entrada_mercadorias.py, Lote P3),
que valida nome, aba, as 20 colunas esperadas e o formato dos valores dos KPIs.
Este aqui e o contrario: nao sabe nada sobre significado, so devolve a
ESTRUTURA (colunas por posicao + linhas cruas) pro perfil deterministico
descrever. Serve as 8 familias que ainda nao tem semantica aprovada -- e por
isso NENHUMA soma sai daqui: quem decide o que pode ser somado e o catalogo
semantico, no perfil (perfil_dados.py).

Tres coisas que o parser do upload manual nao faz e estas fontes exigem
(obstaculos 1 e 2 do docs/FONTES_DATAHUB.md), resolvidas aqui:

- **linha de cabecalho variavel** (1, 2, 3, 5 ou 6): vem da familia
  (nuvem_datahub._FAMILIAS, conferido arquivo por arquivo), pode ser informada
  a mao e, em familia desconhecida, e detectada -- sempre declarando de onde
  veio, pra ninguem confundir palpite com fato.
- **rotulo de coluna repetido** (`EMB` duas vezes): a identidade da coluna e a
  POSICAO; o nome e so rotulo. Nada de dicionario nome->indice aqui.
- **arquivo partido** (`_f1`/`_f2`/`_f3`): cada parte e um arquivo proprio; a
  selecao de varios arquivos numa sessao e o que junta as partes (a
  concatenacao semantica, com o obstaculo do `DADOS_GERAIS` quebrado, nao e
  deste bloco).

Guardas de seguranca (as mesmas do P3, nao relaxadas): so item_id que apareceu
numa sincronizacao (lista de permissao em inventario_datahub), so .xlsx, e
limite de tamanho no download. openpyxl com read_only/data_only nao executa
macro por construcao.
"""

import io
import os

import openpyxl

from . import graph_datahub, inventario_datahub, nuvem_datahub

# Ate onde procurar o cabecalho quando a familia e desconhecida. A maior linha
# de cabecalho conhecida e 6 (SAIDA_MERCADORIAS); 10 da folga sem varrer o
# arquivo inteiro.
_LINHAS_BUSCA_CABECALHO = 10

# Aba preferida: todos os exports do DataHub sao do WMS SLIN e usam esta aba
# (docs/FONTES_DATAHUB.md). Se nao existir, cai na primeira aba do arquivo --
# declarando qual foi usada.
_ABA_PREFERIDA = "SLIN"


class LeituraDatahubError(Exception):
    """Erro de validacao do item/arquivo/aba/cabecalho -- mensagem sempre clara
    pro chamador (endpoint traduz pra HTTP 400)."""


def _limite_bytes() -> int:
    """Mesmo limite do upload manual e do leitor do P3 (UPLOAD_MAX_MB, default
    50 MB) -- sem variavel nova."""
    return int(os.environ.get("UPLOAD_MAX_MB", "50")) * 1024 * 1024


def _arquivo_do_inventario(item_id: str) -> dict:
    if not inventario_datahub.status().get("resumo"):
        raise LeituraDatahubError(
            "nenhuma sincronizacao do DataHub ainda -- clique em 'Sincronizar agora' primeiro"
        )
    arquivo = inventario_datahub.arquivo_por_item_id(item_id)
    if arquivo is None:
        raise LeituraDatahubError("item_id nao encontrado na ultima sincronizacao do DataHub")
    return arquivo


def _escolher_aba(wb):
    if _ABA_PREFERIDA in wb.sheetnames:
        return wb[_ABA_PREFERIDA], _ABA_PREFERIDA
    if not wb.sheetnames:
        raise LeituraDatahubError("arquivo sem nenhuma aba")
    nome = wb.sheetnames[0]
    return wb[nome], nome


def _preenchidas(linha) -> int:
    return sum(1 for v in linha if v is not None and str(v).strip() != "")


def _detectar_linha_cabecalho(linhas: list[list]) -> int:
    """Melhor esforco pra familia desconhecida: a linha com mais celulas
    preenchidas entre as primeiras. Empate fica com a primeira (o cabecalho
    vem antes dos dados). Devolve 1-based; sem nenhuma linha preenchida,
    devolve 1 -- o perfil declara que foi detectado, nunca finge certeza."""
    melhor_indice, melhor_contagem = 0, 0
    for i, linha in enumerate(linhas):
        contagem = _preenchidas(linha)
        if contagem > melhor_contagem:
            melhor_indice, melhor_contagem = i, contagem
    return melhor_indice + 1


def _colunas(linha_cabecalho: list) -> list[dict]:
    """Identidade pela POSICAO (1-based); nome e so rotulo, pode repetir ou
    vir vazio."""
    colunas = []
    for i, valor in enumerate(linha_cabecalho):
        nome = str(valor).strip() if valor is not None else ""
        colunas.append({"posicao": i + 1, "nome": nome or f"(coluna {i + 1})"})
    # corta o rabo de colunas vazias que o Excel costuma deixar
    while colunas and colunas[-1]["nome"].startswith("(coluna "):
        colunas.pop()
    if not colunas:
        raise LeituraDatahubError("linha de cabecalho vazia -- nenhuma coluna encontrada")
    return colunas


def ler_estrutura(item_id: str, max_linhas: int, linha_cabecalho: int | None = None) -> dict:
    """Baixa e le a estrutura de um arquivo ja sincronizado.

    max_linhas: teto de linhas de dado devolvidas (o resto e contado, nao
    devolvido -- `truncado` e `linhas_lidas` dizem o que aconteceu).
    linha_cabecalho: 1-based, sobrepoe a da familia (quando o usuario corrige
    na tela).
    """
    arquivo = _arquivo_do_inventario(item_id)
    nome = arquivo["nome"]
    if not nome.lower().endswith(".xlsx"):
        raise LeituraDatahubError(
            f"perfil so le .xlsx -- este arquivo nao e planilha: {nome}"
        )

    definicao = nuvem_datahub.definicao_do_arquivo(nome)
    if linha_cabecalho is not None:
        if linha_cabecalho < 1:
            raise LeituraDatahubError("linha de cabecalho tem que ser 1 ou maior")
        origem_cabecalho = "informada"
    elif definicao.get("linha_cabecalho"):
        linha_cabecalho = definicao["linha_cabecalho"]
        origem_cabecalho = "familia"
    else:
        origem_cabecalho = "detectada"

    conteudo = graph_datahub.baixar_item(item_id, limite_bytes=_limite_bytes())
    try:
        wb = openpyxl.load_workbook(io.BytesIO(conteudo), read_only=True, data_only=True)
    except Exception as exc:
        raise LeituraDatahubError("arquivo nao e um .xlsx valido ou esta corrompido") from exc

    try:
        ws, aba = _escolher_aba(wb)
        abas = list(wb.sheetnames)
        iterador = ws.iter_rows(values_only=True)

        if origem_cabecalho == "detectada":
            inicio = [list(l) for _, l in zip(range(_LINHAS_BUSCA_CABECALHO), iterador)]
            linha_cabecalho = _detectar_linha_cabecalho(inicio)
            cabecalho = inicio[linha_cabecalho - 1] if inicio else []
            # o que sobrou do bloco inicial ja e dado
            pendentes = inicio[linha_cabecalho:]
        else:
            pendentes = []
            cabecalho = None
            for numero, linha in enumerate(iterador, start=1):
                if numero == linha_cabecalho:
                    cabecalho = list(linha)
                    break
            if cabecalho is None:
                raise LeituraDatahubError(
                    f"arquivo tem menos de {linha_cabecalho} linha(s) -- cabecalho "
                    f"esperado na linha {linha_cabecalho}"
                )

        colunas = _colunas(cabecalho)
        largura = len(colunas)

        linhas: list[list] = []
        lidas = 0
        for linha in [*pendentes, *iterador]:
            if linha is None or all(v is None for v in linha):
                continue
            lidas += 1
            if len(linhas) < max_linhas:
                valores = list(linha[:largura])
                valores += [None] * (largura - len(valores))
                linhas.append(valores)
    finally:
        wb.close()

    filial, competencia = nuvem_datahub.filial_competencia_do_arquivo(nome)
    return {
        "item_id": item_id,
        "arquivo": nome,
        "caminho": arquivo.get("caminho"),
        "web_url": arquivo.get("web_url"),
        "tamanho": arquivo.get("tamanho"),
        "modificado_em": arquivo.get("modificado_em"),
        "familia": definicao["familia"],
        "area": definicao["area"],
        "estado_familia": definicao["estado"],
        "filial": filial,
        "competencia": competencia,
        "aba": aba,
        "abas": abas,
        "linha_cabecalho": linha_cabecalho,
        "origem_linha_cabecalho": origem_cabecalho,
        "colunas": colunas,
        "linhas": linhas,
        "linhas_lidas": lidas,
        "truncado": lidas > len(linhas),
    }
